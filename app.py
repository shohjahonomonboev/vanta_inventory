# app.py — Vanta Inventory (FINAL BEAST, Orin-patched)

# ── Stdlib
import os, io, time, json, zipfile, logging, sys, re, hmac, secrets
from decimal import Decimal, InvalidOperation
from urllib.parse import urlparse
from datetime import datetime, date, timedelta

# Make sqlite accept Decimal transparently
try:
    import sqlite3
    sqlite3.register_adapter(Decimal, float)
except Exception:
    pass

# ── Third-party
import requests
from werkzeug.exceptions import HTTPException
from werkzeug.middleware.proxy_fix import ProxyFix
from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, send_file, jsonify, g
)

# Optional Babel for pretty money format
try:
    from babel.numbers import format_currency as _format_currency
except Exception:  # fallback if Babel not installed
    def _format_currency(value, currency, locale=None):
        try:
            return f"{float(value or 0):,.2f} {currency}"
        except Exception:
            return f"{value} {currency}"

# ── Local modules
import database_sqlalchemy as db
from database_sqlalchemy import ensure_schema
from i18n import t  # translation helper

# ── App config
OFFLINE = os.getenv("OFFLINE", "0").lower() in ("1", "true", "yes")
ENV = os.getenv("ENV", "dev").lower()
IS_DEV = (ENV == "dev")

def _parse_date(s: str):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


# Flash helpers
def flash_success(msg): flash(msg, "success")
def flash_error(msg):   flash(msg, "error")
def flash_info(msg):    flash(msg, "info")
def flash_warn(msg):    flash(msg, "warning")

# =========================
# App bootstrap + env
# =========================
app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "dev-only-change-me")

app.config.update(
    DEBUG=IS_DEV,
    TEMPLATES_AUTO_RELOAD=IS_DEV,
    SEND_FILE_MAX_AGE_DEFAULT=(0 if IS_DEV else 31536000),
    SESSION_COOKIE_SECURE=not IS_DEV,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_HTTPONLY=True,
    PERMANENT_SESSION_LIFETIME=timedelta(days=7),
)

# app.py (top-level, near other config)
import os
APP_VERSION = os.getenv("APP_VERSION", "dev")
GIT_SHA = os.getenv("RENDER_GIT_COMMIT") or os.getenv("SOURCE_VERSION") or os.getenv("GIT_SHA", "")

@app.route("/__version__")
def __version__():
    return {"app_version": APP_VERSION, "git_sha": GIT_SHA[:8]}


# Behind-proxy headers (prod only)
if not IS_DEV:
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1)

# =========================
# Logging
# =========================
handler = logging.StreamHandler(sys.stdout)
handler.setLevel(logging.INFO)
app.logger.addHandler(handler)
app.logger.setLevel(logging.INFO)

# =========================
# Security headers
# =========================
@app.after_request
def secure_headers(resp):
    resp.headers.setdefault("X-Content-Type-Options", "nosniff")
    resp.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    resp.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    resp.headers.setdefault(
        "Permissions-Policy",
        "geolocation=(), microphone=(), camera=(), interest-cohort=()"
    )
    return resp

@app.route("/__debug/check_admin_pw", methods=["GET", "POST"])
def __debug_check_admin_pw():
    if not IS_DEV or not is_logged_in():
        return "Forbidden", 403

    if request.method == "POST":
        sent = (
            (request.form.get("pw"))
            or ((request.is_json and request.json.get("pw")) or "")
            or ""
        )
        sent = str(sent).strip()
        ok = hmac.compare_digest(ADMIN_ACTION_PASSWORD, sent)
        return jsonify({
            "ok": bool(ok),
            "entered_len": len(sent or ""),
            "configured_len": len(ADMIN_ACTION_PASSWORD or ""),
            "gate_on": bool(ADMIN_ACTION_PASSWORD)
        })

    return """
      <form method="post">
        <input type="password" name="pw" placeholder="Enter admin password" />
        <button type="submit">Test</button>
      </form>
    """, 200


@app.get("/__debug/admin_gate_status")
def __debug_admin_gate_status():
    if not is_logged_in():
        return "Login required", 401
    st, remain = _lock_state()
    src = (
        "ADMIN_ACTION_PASSWORD" if os.getenv("ADMIN_ACTION_PASSWORD")
        else "ORIN_ADMIN_PASSWORD" if os.getenv("ORIN_ADMIN_PASSWORD")
        else "ORIN_PASSWORD" if os.getenv("ORIN_PASSWORD")
        else "DEFAULT:changeme"
    )
    return jsonify({
        "gate_on": bool(ADMIN_ACTION_PASSWORD),
        "configured_len": len(ADMIN_ACTION_PASSWORD or ""),
        "source_env": src,
        "locked_remaining_seconds": remain,
        "fails": st.get("fails", 0),
        "env_samples": {
            "ADMIN_ACTION_PASSWORD": bool(os.getenv("ADMIN_ACTION_PASSWORD")),
            "ORIN_ADMIN_PASSWORD":  bool(os.getenv("ORIN_ADMIN_PASSWORD")),
            "ORIN_PASSWORD":        bool(os.getenv("ORIN_PASSWORD")),
        }
    })


# =========================
# Global prod error page
# =========================
if not IS_DEV:
    @app.errorhandler(Exception)
    def handle_error(e):
        if isinstance(e, HTTPException):
            return e
        app.logger.exception("Unhandled exception")
        return "Internal Server Error", 500

# =========================
# Ensure DB schema at boot
# =========================
with app.app_context():
    try:
        ensure_schema()
        # Ensure admin_logs table
        try:
            if db.DATABASE_URL.startswith("sqlite"):
                db.db_exec("""
                    CREATE TABLE IF NOT EXISTS admin_logs (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        action TEXT,
                        item TEXT,
                        success INTEGER,
                        ts TEXT
                    )
                """)
            else:
                db.db_exec("""
                    CREATE TABLE IF NOT EXISTS admin_logs (
                        id SERIAL PRIMARY KEY,
                        action TEXT,
                        item TEXT,
                        success INTEGER,
                        ts TEXT
                    )
                """)
            app.logger.info("✅ admin_logs table ensured.")
        except Exception:
            app.logger.exception("⚠️ Failed to ensure admin_logs table")
        app.logger.info("✅ Database schema ensured at startup.")
    except Exception:
        app.logger.exception("⚠️ Failed to ensure schema")

# =========================
# Version / Health
# =========================
def get_version():
    val = None
    try:
        val = open("VERSION", encoding="utf-8").read().strip()
    except Exception:
        pass
    if not val or len(val) < 3:
        val = os.getenv("APP_VERSION", "").strip()
    if not val or len(val) < 3:
        val = "v1.0.0"
    return val

@app.context_processor
def inject_version():
    return {"APP_VERSION": get_version()}

@app.get("/__health")
def __health():
    return jsonify(status="ok", version=get_version()), 200

# =========================
# CSRF
# =========================
def _csrf_token():
    tok = session.get("_csrf")
    if not tok:
        tok = secrets.token_urlsafe(32)
        session["_csrf"] = tok
    return tok

@app.context_processor
def _inject_csrf():
    return {"CSRF_TOKEN": _csrf_token()}

def check_csrf() -> bool:
    expected = session.get("_csrf", "")
    sent = request.form.get("_csrf") or request.headers.get("X-CSRF-Token") or ""
    ok = bool(expected) and hmac.compare_digest(expected, sent)
    if not ok:
        flash("Security check failed. Please retry.", "error")
        return False
    return True

@app.before_request
def _csrf_protect():
    if request.method in ("POST", "PUT", "PATCH", "DELETE"):
        expected = session.get("_csrf", "")
        sent = request.form.get("_csrf") or request.headers.get("X-CSRF-Token") or ""
        ok = bool(expected) and hmac.compare_digest(expected, sent)
        if not ok:
            app.logger.warning("CSRF blocked: path=%s", request.path)
            wants_json = request.is_json or "application/json" in request.headers.get("Accept", "")
            if wants_json:
                return jsonify({"error": "CSRF validation failed"}), 400
            flash("Security check failed. Please retry.", "error")
            return redirect(request.referrer or url_for("index"))

# =========================
# Currency / i18n helpers
# =========================
BASE_CCY = "UZS"          # DB-stored currency
DEFAULT_LANG = "en"
DEFAULT_CURR = "USD"
_SUPPORTED = {"USD", "AED", "UZS"}

def get_lang():
    return session.get("LANG", DEFAULT_LANG)

def get_curr():
    return session.get("CURR", DEFAULT_CURR)

def fmt_money(value, curr=None, lang=None):
    curr = curr or get_curr()
    lang = lang or get_lang()
    locale = "en_US" if lang == "en" else "uz_UZ"
    try:
        return _format_currency(float(value or 0), curr, locale=locale)
    except Exception:
        try:
            return f"{float(value):.2f} {curr}"
        except Exception:
            return f"{value} {curr}"

# FX cache: 1 USD = rate[currency]
_FX_CACHE = {"rates": None, "ts": 0}

def _fetch_usd_rates():
    if OFFLINE:
        return {"USD": 1.0, "AED": 3.6725, "UZS": 12600.0}

    now = time.time()
    if _FX_CACHE["rates"] and (now - _FX_CACHE["ts"]) < 3600:
        return _FX_CACHE["rates"]

    providers = [
        ("exchangerate.host", lambda: requests.get(
            "https://api.exchangerate.host/latest",
            params={"base": "USD", "symbols": "USD,AED,UZS"},
            headers={"User-Agent": "VantaInventory/1.0"}, timeout=8
        ).json().get("rates", {})),
        ("open.er-api.com", lambda: requests.get(
            "https://open.er-api.com/v6/latest/USD",
            headers={"User-Agent": "VantaInventory/1.0"}, timeout=8
        ).json().get("rates", {})),
        ("frankfurter.app", lambda: requests.get(
            "https://api.frankfurter.app/latest",
            params={"from": "USD", "to": "AED,USD,UZS"},
            headers={"User-Agent": "VantaInventory/1.0"}, timeout=8
        ).json().get("rates", {})),
    ]

    rates = None
    for _, fn in providers:
        try:
            r = fn() or {}
            out = {
                "USD": float(r.get("USD", 1.0)),
                "AED": float(r.get("AED")) if r.get("AED") is not None else None,
                "UZS": float(r.get("UZS")) if r.get("UZS") is not None else None,
            }
            if out["AED"] is None and "AED" in r and r["AED"]:
                out["AED"] = float(r["AED"])
            if out["UZS"] is None and "UZS" in r and r["UZS"]:
                out["UZS"] = float(r["UZS"])
            if out["AED"] is not None:
                if out["UZS"] is None:
                    prev = _FX_CACHE.get("rates") or {}
                    out["UZS"] = float(prev.get("UZS", 12600.0))
                rates = out
                break
        except Exception:
            continue

    if not rates:
        rates = {"USD": 1.0, "AED": 3.6725, "UZS": 12600.0}

    _FX_CACHE.update({"rates": rates, "ts": time.time()})
    return rates

def _derive_rates_from_usd(base):
    R = _fetch_usd_rates()
    base = (base or "USD").upper()
    if base not in R:
        base = "USD"
    out = {}
    for x in _SUPPORTED:
        if x == base:
            out[x] = 1.0
        else:
            try:
                out[x] = float(R[x]) / float(R[base])
            except Exception:
                out[x] = 0.0
    return base, out

def convert_amount(value, from_curr=None, to_curr=None):
    try:
        v = float(value or 0)
    except Exception:
        return 0.0
    from_curr = (from_curr or BASE_CCY).upper()
    to_curr   = (to_curr or get_curr()).upper()
    if from_curr == to_curr:
        return v
    R = _fetch_usd_rates()
    if from_curr not in R or to_curr not in R:
        return v
    amount_usd = v / float(R[from_curr])
    return amount_usd * float(R[to_curr])

def fmt_money_auto(value, from_curr=None):
    return fmt_money(convert_amount(value, from_curr=from_curr, to_curr=get_curr()))

# Money helpers (for “1,234” inputs)
def money_plain(v, decimals=0):
    try:
        v = float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return v
    return f"{v:,.{decimals}f}"

def parse_money(s):
    try:
        return float(str(s).replace(",", "").strip() or 0)
    except ValueError:
        return 0.0

# =========================
# Auth (hardened)
# =========================
def _load_admins():
    """
    Load admins from env:
      ADMIN_USERS_JSON='{"vanta":"new2025"}'
      -or-
      ADMIN_USERS='vanta:new2025'
    Fallback includes ONLY vanta:new2025.
    Keys stored casefold() for case-insensitive matching.
    """
    raw_json = os.environ.get("ADMIN_USERS_JSON", "").strip()
    if raw_json:
        try:
            data = json.loads(raw_json)
            return { (k or "").casefold(): str(v) for k, v in data.items() }
        except Exception:
            pass

    raw_pairs = os.environ.get("ADMIN_USERS", "").strip()
    if raw_pairs:
        try:
            pairs = dict(pair.split(":", 1) for pair in raw_pairs.split(","))
            return { (k or "").casefold(): str(v) for k, v in pairs.items() }
        except Exception:
            pass

    # 🔒 Fallback: only Vanta, only new2025
    return {"vanta": "new2025"}


ADMIN_USERS = _load_admins()
ADMIN_ADMINS = set(ADMIN_USERS.keys())  # allow-list

def is_logged_in():
    return bool(session.get("logged_in"))

def is_admin_user():
    return (session.get("user") or "").casefold() in ADMIN_ADMINS

# --- Admin action password / lockout (SINGLE SOURCE OF TRUTH) ---
ADMIN_ACTION_PASSWORD = (
    os.getenv("ADMIN_ACTION_PASSWORD")
    or os.getenv("ORIN_ADMIN_PASSWORD")
    or os.getenv("ORIN_PASSWORD")
    or "changeme"  # not empty, so the gate is ON by default
).strip()

# ▼▼▼ Added: admin-gate helpers that were referenced but not defined ▼▼▼
def _lock_state():
    st = session.get("_admin_gate")
    if not isinstance(st, dict):
        st = {"fails": 0, "lock_until": 0}
    now = int(time.time())
    remain = max(0, int(st.get("lock_until", 0)) - now)
    return st, remain

def _lock_fail():
    st, remain = _lock_state()
    now = int(time.time())
    st["fails"] = int(st.get("fails", 0)) + 1
    # progressive lock: after 3+ fails, lock for 60s * fails (cap 10min)
    if st["fails"] >= 3:
        st["lock_until"] = now + min(600, 60 * st["fails"])
    session["_admin_gate"] = st

def _lock_clear():
    session["_admin_gate"] = {"fails": 0, "lock_until": 0}

def _admin_pw_ok(sent: str) -> bool:
    # If gate password is empty, gate is OFF
    if not ADMIN_ACTION_PASSWORD:
        return True
    try:
        return hmac.compare_digest(ADMIN_ACTION_PASSWORD, sent or "")
    except Exception:
        return False
# ▲▲▲ End helpers ▲▲▲

def require_admin_action_pw() -> bool:
    # lockout check
    st, remain = _lock_state()
    if remain > 0:
        flash(f"Admin actions locked. Try again in {remain}s.", "error")
        app.logger.warning(f"[admin_gate] locked: remaining={remain}s fails={st.get('fails')}")
        return False

    # collect candidate sent password
    sent = (
        request.form.get("admin_password")
        or request.form.get("admin_pw")
        or request.form.get("password")
        or request.headers.get("X-Admin-Password")
        or request.args.get("admin_password")    # last resort (debug)
        or request.values.get("apw")             # last resort (debug)
        or ""
    )
    sent = str(sent).strip()

    # lookup current user's stored login password (case-insensitive key)
    user_key = (session.get("user") or "").casefold()
    user_pw  = ADMIN_USERS.get(user_key, "")

    # log sources (never the actual password)
    app.logger.info("[admin_gate] sources: %s", {
        "form_admin_password": "admin_password" in request.form,
        "form_admin_pw": "admin_pw" in request.form,
        "form_password": "password" in request.form,
        "header": bool(request.headers.get("X-Admin-Password")),
        "query": ("admin_password" in request.args) or ("apw" in request.args),
        "len": len(sent or ""),
        "user": user_key,
        "has_user_pw": bool(user_pw),
    })

    # accept if matches either the env gate OR the logged-in user's login password
    ok = _admin_pw_ok(sent)  # accept ONLY the env gate

    if not ok:
        _lock_fail()
        st2, rem2 = _lock_state()
        app.logger.warning(f"[admin_gate] bad password (fails={st2['fails']} remain={rem2}s)")
        flash("Admin password incorrect.", "error")
        return False

    _lock_clear()
    return True


def _admin_log(action: str, item: str, success: int):
    try:
        db.db_exec(
            "INSERT INTO admin_logs (action, item, success, ts) VALUES (:a,:i,:s,:t)",
            {"a": action, "i": item, "s": int(success), "t": datetime.utcnow().isoformat()}
        )
    except Exception:
        app.logger.exception("[admin_log] failed")

@app.get("/admin/lockout")
def admin_lockout():
    _, remain = _lock_state()
    return jsonify({"remaining": remain})

# Quick unlock for your session (requires login)
@app.get("/admin/unlock")
def admin_unlock():
    if not is_logged_in():
        return redirect(url_for("login"))
    _lock_clear()
    flash("Admin lock cleared for this session.", "info")
    return redirect(url_for("index"))

# Log gate status at boot (no secret leak)
app.logger.info("[config] Admin gate: %s (len=%d)", "ON" if ADMIN_ACTION_PASSWORD else "OFF", len(ADMIN_ACTION_PASSWORD))

# =========================
# Jinja filters / context
# =========================
@app.before_request
def _inject_currency():
    g.CURRENCY = get_curr()

@app.template_filter("ccy")
def ccy(amount):
    # DB currency (UZS) → UI currency
    return convert_amount(amount, from_curr=BASE_CCY, to_curr=get_curr())

@app.template_filter("fmtmoney")
def fmtmoney(amount):
    return fmt_money(amount, curr=get_curr(), lang=get_lang())

@app.template_filter("comma")
def comma(x):
    try:
        return f"{int(x):,}"
    except Exception:
        return x

@app.template_filter("moneyfmt")
def moneyfmt(x, currency=""):
    try:
        return f"{int(float(x)):,} {currency}" if currency else f"{int(float(x)):,}"
    except Exception:
        return x

@app.template_filter("money")
def money_filter(v):
    try:
        return money_plain(v, decimals=0)
    except Exception:
        return v

@app.context_processor
def inject_helpers():
    lang = get_lang()
    return {
        "fmt_money": fmt_money,
        "fmt_money_auto": fmt_money_auto,
        "convert": convert_amount,
        "USER_LANG": lang,
        "USER_CURR": get_curr(),
        "CURRENCY": get_curr(),
        "t": lambda key, _lang=None: t(key, lang if _lang is None else _lang),
    }

# =========================
# Data access
# =========================
def get_inventory():
    rows = db.db_all(
        """
        SELECT
            id,
            name,
            buying_price,
            selling_price,
            quantity,
            CASE
              WHEN profit IS NULL OR profit = 0
                THEN (selling_price - buying_price) * quantity
              ELSE profit
            END AS profit,
            COALESCE(currency, :c) AS currency
        FROM inventory
        ORDER BY id
    """,
        {"c": BASE_CCY},
    )
    return [tuple(r) for r in rows]

# =========================
# Pages / Routes
# =========================
@app.route("/")
def index():
    if not is_logged_in():
        return redirect(url_for("login"))

    # ----- query params -----
    search_query    = (request.values.get("search", "") or "").strip().lower()
    selected_filter = request.values.get("filter", "")
    sort_by         = request.args.get("sort_by", "id")
    direction       = request.args.get("direction", "asc")
    sort_param      = request.args.get("sort")

    # ----- Unified date window (engine-agnostic, inclusive end by using exclusive next-day) -----
    start_str = (request.args.get("from") or "").strip()
    end_str   = (request.args.get("to")   or "").strip()

    # Helper to format YYYY-MM-DD safely
    def _coerce_date_str(s: str) -> str:
        s = (s or "").strip()
        return s[:10] if len(s) >= 10 else ""

    from datetime import datetime, timedelta

    today = datetime.now().date()

    # build start date
    if _coerce_date_str(start_str):
        start_date = datetime.strptime(_coerce_date_str(start_str), "%Y-%m-%d").date()
    else:
        start_date = today

    # build end date (inclusive in UI, exclusive in SQL by adding 1 day)
    if _coerce_date_str(end_str):
        end_date = datetime.strptime(_coerce_date_str(end_str), "%Y-%m-%d").date()
    else:
        end_date = today

    # ensure order (if user inverted)
    if end_date < start_date:
        start_date, end_date = end_date, start_date

    start_ts = f"{start_date.isoformat()} 00:00:00"
    end_exclusive_ts = f"{(end_date + timedelta(days=1)).isoformat()} 00:00:00"

    # Single WHERE for both engines (string timestamps work in SQLite and Postgres)
    where  = "s.sold_at >= :start AND s.sold_at < :end"
    params = {"start": start_ts, "end": end_exclusive_ts}

    # (optional, but super helpful)
    app.logger.info("[date-window] start=%s end_exclusive=%s", start_ts, end_exclusive_ts)

    using_sqlite = db.DATABASE_URL.startswith("sqlite")


    if using_sqlite:
        where_clauses, params = [], {}
        if start_str:
            where_clauses.append("s.sold_at >= :start")
            params["start"] = f"{start_str} 00:00:00"
        if end_str:
            where_clauses.append("s.sold_at < DATETIME(:end, '+1 day')")  # inclusive
            params["end"] = f"{end_str} 00:00:00"
        if not where_clauses:
            where = "s.sold_at >= DATETIME('now','start of day') AND s.sold_at < DATETIME('now','start of day','+1 day')"
        else:
            where = " AND ".join(where_clauses)
    else:
        where_clauses, params = [], {}
        if start_str:
            where_clauses.append("s.sold_at >= DATE(:start)")
            params["start"] = start_str
        if end_str:
            where_clauses.append("s.sold_at < DATE(:end) + INTERVAL '1 day'")  # inclusive
            params["end"] = end_str
        if not where_clauses:
            where = "s.sold_at >= CURRENT_DATE AND s.sold_at < CURRENT_DATE + INTERVAL '1 day'"
        else:
            where = " AND ".join(where_clauses)

    # map short sort tokens for inventory table (client menu)
    if sort_param:
        mapping = {
            "name_asc":  ("name", "asc"),
            "name_desc": ("name", "desc"),
            "qty_asc":   ("quantity", "asc"),
            "qty_desc":  ("quantity", "desc"),
            "price_asc": ("price", "asc"),
            "price_desc":("price", "desc"),
        }
        sort_by, direction = mapping.get(sort_param, (sort_by, direction))

    # ----- data -----
    inventory = get_inventory()  # (id, name, buy, sell, qty, profit, currency)

    # ---- KPIs: (REUSED WHERE) ----
    row = db.db_one(
        f"""
        SELECT
          COALESCE(SUM(s.qty * s.sell_price), 0),
          COALESCE(SUM(s.profit), 0)
        FROM sales s
        WHERE {where}
        """,
        params
    )
    today_revenue = row[0] if row else 0
    today_profit  = row[1] if row else 0

    # ---- Filtering (in-memory) ----
    if search_query:
        inventory = [it for it in inventory if search_query in (it[1] or "").lower()]
    if selected_filter == "low_stock":
        inventory = [it for it in inventory if (it[4] or 0) <= 5]
    elif selected_filter == "high_profit":
        inventory = [it for it in inventory if (it[5] or 0) >= 100]

    # ---- Sorting (in-memory) ----
    sort_map = {"name": 1, "quantity": 4, "profit": 5, "price": 3}
    if sort_by in sort_map:
        idx = sort_map[sort_by]
        inventory = sorted(
            inventory,
            key=lambda x: (x[idx] is None, x[idx] if x[idx] is not None else 0),
            reverse=(direction == "desc"),
        )

    # ---- Totals ----
    def nz_dec(x): return x if x is not None else Decimal(0)
    def nz_int(x): return x if x is not None else 0

    total_quantity = sum(nz_int(it[4]) for it in inventory)
    total_profit   = sum(nz_dec(it[5]) for it in inventory)

    # ---- Top/Low lists ----
    top_profit_items = sorted(inventory, key=lambda x: (x[5] is None, x[5]), reverse=True)[:5]
    low_stock_items  = sorted(inventory, key=lambda x: (x[4] is None, x[4]))[:5]

    # ---- Last 7 days revenue (chart) ----
    if using_sqlite:
        rows = db.db_all(
            """
            WITH days AS (
              SELECT DATE('now','localtime','-6 day') AS d
              UNION ALL SELECT DATE('now','localtime','-5 day')
              UNION ALL SELECT DATE('now','localtime','-4 day')
              UNION ALL SELECT DATE('now','localtime','-3 day')
              UNION ALL SELECT DATE('now','localtime','-2 day')
              UNION ALL SELECT DATE('now','localtime','-1 day')
              UNION ALL SELECT DATE('now','localtime')
            )
            SELECT d AS day,
                   COALESCE((SELECT SUM(qty*sell_price)
                             FROM sales s
                             WHERE DATE(s.sold_at)=d),0) AS revenue
            FROM days
            """
        )
    else:
        rows = db.db_all(
            """
            WITH days AS (
              SELECT generate_series(current_date - interval '6 day', current_date, interval '1 day')::date AS d
            )
            SELECT d AS day,
                   COALESCE((SELECT SUM(qty*sell_price) FROM sales s WHERE DATE(s.sold_at)=d),0) AS revenue
            FROM days
            ORDER BY d
            """
        )

    sales_labels = [str(r[0]) for r in rows]
    sales_values = [float(r[1] or 0) for r in rows]

    # ---- Stock tracker (legacy arrays) ----
    stock_labels = [it[1] for it in inventory]
    stock_values = [nz_int(it[4]) for it in inventory]

    # ---- ORIN ADD: Sold list + pagination (REUSED WHERE) ----
    page = max(int(request.args.get("page", 1) or 1), 1)
    per_page = 50
    offset = (page - 1) * per_page

    sales_today = db.db_all(
        f"""
        SELECT s.id, s.item_id, i.name, s.qty, s.sell_price, s.profit, s.sold_at
        FROM sales s
        JOIN inventory i ON i.id = s.item_id
        WHERE {where}
        ORDER BY s.sold_at DESC
        LIMIT :limit OFFSET :offset
        """,
        {**params, "limit": per_page, "offset": offset}
    )

    # ---- ORIN ADD: total count & pages ----
    total_count = db.db_one(
        f"""
        SELECT COUNT(*)
        FROM sales s
        JOIN inventory i ON i.id = s.item_id
        WHERE {where}
        """,
        params
    )[0]
    total_pages = max((total_count + per_page - 1) // per_page, 1)

    # ---- FX footer ----
    _base, _rates = _derive_rates_from_usd("USD")
    usd_to_aed = _rates.get("AED", 0)
    usd_to_uzs = _rates.get("UZS", 0)

    # ---- Render ----
    ctx = {
        "selected_from": start_str,
        "selected_to": end_str,
        "inventory": inventory,
        "search_query": search_query,
        "sort_by": sort_by,
        "direction": direction,
        "selected_filter": selected_filter,
        "total_quantity": total_quantity,
        "total_profit": total_profit,
        "top_profit_labels": [it[1] for it in top_profit_items],
        "top_profit_values": [float(nz_dec(it[5])) for it in top_profit_items],
        "low_stock_labels": [it[1] for it in low_stock_items],
        "low_stock_values": [nz_int(it[4]) for it in low_stock_items],
        "sales_labels": sales_labels,
        "sales_values": sales_values,
        "stock_labels": stock_labels,
        "stock_values": stock_values,
        "sales_today": sales_today,
        "today_revenue": today_revenue,
        "today_profit": today_profit,
        "usd_to_aed": usd_to_aed,
        "usd_to_uzs": usd_to_uzs,
        "page": page,
        "total_pages": total_pages,
        "total_count": total_count,
    }

    return render_template("index.html", **ctx)

# ➕ Add Item (UPSERT by name)
@app.post("/add")
def add_item():
    if not is_logged_in():
        flash("Please log in.", "warning")
        return redirect(url_for("login"))
    if not check_csrf():
        return redirect(url_for("index"))

    # ---------- Name ----------
    name_raw = (request.form.get("name") or "").strip()
    if not name_raw:
        flash("Name is required.", "error")
        return redirect(url_for("index"))
    name = re.sub(r"\s+", " ", name_raw).strip()[:100]

    # ---------- Quantity ----------
    try:
        quantity = int((request.form.get("quantity") or "").strip())
    except Exception:
        flash("Quantity must be a number.", "error")
        return redirect(url_for("index"))
    if quantity <= 0:
        flash("Quantity must be greater than 0.", "error")
        return redirect(url_for("index"))

    # ---------- Prices (UI currency) ----------
    try:
        bp_ui = Decimal(str(parse_money(request.form.get("buying_price") or "0")))
        sp_ui = Decimal(str(parse_money(request.form.get("selling_price") or "0")))
    except (InvalidOperation, Exception):
        flash("Prices must be numeric.", "error")
        return redirect(url_for("index"))
    if bp_ui < 0 or sp_ui < 0:
        flash("Prices must be non-negative.", "error")
        return redirect(url_for("index"))
    if sp_ui < bp_ui:
        flash("Selling price is below buying price.", "warning")

    # ---------- Convert UI → DB base (UZS) ----------
    try:
        buying_price  = Decimal(str(convert_amount(bp_ui, from_curr=get_curr(), to_curr=BASE_CCY)))
        selling_price = Decimal(str(convert_amount(sp_ui, from_curr=get_curr(), to_curr=BASE_CCY)))
    except Exception as e:
        flash(f"Currency conversion failed: {e}", "error")
        return redirect(url_for("index"))

    # ---------- Cast for SQLite ----------
    bp_db = float(buying_price)
    sp_db = float(selling_price)

    # ---------- UPSERT ----------
    try:
        row = db.db_one("SELECT id, quantity FROM inventory WHERE name=:n", {"n": name})
        if row:
            item_id, existing_qty = row[0], int(row[1] or 0)
            new_qty = existing_qty + quantity
            db.db_exec(
                """
                UPDATE inventory
                   SET buying_price = :bp,
                       selling_price = :sp,
                       quantity      = :q,
                       updated_at    = CURRENT_TIMESTAMP
                 WHERE id = :id
                """,
                {"bp": bp_db, "sp": sp_db, "q": new_qty, "id": item_id},
            )
            flash(f"Updated “{name}”: +{quantity} → {new_qty}.", "success")
        else:
            db.db_exec(
                """
                INSERT INTO inventory (name, buying_price, selling_price, quantity)
                VALUES (:n, :bp, :sp, :q)
                """,
                {"n": name, "bp": bp_db, "sp": sp_db, "q": quantity},
            )
            flash(f"Added “{name}” (qty {quantity}).", "success")
    except Exception as e:
        app.logger.exception("[add_item] failed")
        flash(f"Failed to save item: {e}", "error")

    return redirect(url_for("index"))

# 🛒 Sell Item
@app.post("/sell")
def sell_item():
    if not is_logged_in():
        return redirect(url_for("login"))
    if not check_csrf():
        return redirect(url_for("index"))

    try:
        item_id = int(request.form["item_id"])
        qty     = int(request.form["qty"])
    except Exception:
        flash("Invalid item or quantity.", "error")
        return redirect(url_for("index"))

    raw_price = request.form.get("sell_price") or request.form.get("price") or request.form.get("sell")
    sell_price_ui = parse_money(raw_price)
    sell_price    = convert_amount(sell_price_ui, from_curr=get_curr(), to_curr=BASE_CCY)

    row = db.db_one("SELECT quantity, buying_price FROM inventory WHERE id=:id", {"id": item_id})
    if not row:
        flash("Item not found!", "error")
        return redirect(url_for("index"))

    stock_qty, buy_price = int(row[0] or 0), float(row[1] or 0)
    if qty <= 0:
        flash("Quantity must be greater than 0.", "error")
        return redirect(url_for("index"))
    if qty > stock_qty:
        flash("Not enough stock!", "error")
        return redirect(url_for("index"))

    profit = (sell_price - buy_price) * qty

    db.db_exec(
        """
        INSERT INTO sales (item_id, qty, sell_price, profit, sold_at)
        VALUES (:id, :q, :sp, :pf, CURRENT_TIMESTAMP)
        """,
        {"id": item_id, "q": qty, "sp": sell_price, "pf": profit},
    )
    db.db_exec(
        "UPDATE inventory SET quantity = quantity - :q WHERE id = :id",
        {"q": qty, "id": item_id},
    )

    flash(f"Sold {qty} unit(s). Profit: {fmt_money(profit)}", "success")
    return redirect(url_for("index"))

# ↩️ Return sale (restock + remove record) — PASSWORD PROTECTED
@app.post("/sales/<int:sale_id>/return")
def return_sale(sale_id):
    if not is_logged_in(): return redirect(url_for("login"))
    if not check_csrf():   return redirect(url_for("index"))
    if not require_admin_action_pw(): return redirect(url_for("index"))

    try:
        sale = db.db_one("SELECT id, item_id, qty FROM sales WHERE id=:id", {"id": sale_id})
        if not sale:
            flash("Sale not found.", "error"); return redirect(url_for("index"))

        item_row = db.db_one("SELECT name FROM inventory WHERE id=:id", {"id": sale[1]})
        item_name = item_row[0] if item_row else "unknown"

        qty = int(sale[2] or 0)
        if qty <= 0:
            db.db_exec("DELETE FROM sales WHERE id=:id", {"id": sale_id})
            _admin_log("return", item_name, 1)
            flash("Sale removed (no quantity to return).", "warning")
            return redirect(url_for("index"))

        inv = db.db_one("SELECT id FROM inventory WHERE id=:id", {"id": sale[1]})
        if not inv:
            flash("Linked inventory item not found.", "error"); return redirect(url_for("index"))

        db.db_exec("UPDATE inventory SET quantity = quantity + :q WHERE id = :id", {"q": qty, "id": sale[1]})
        db.db_exec("DELETE FROM sales WHERE id=:id", {"id": sale_id})
        _admin_log("return", item_name, 1)
        flash("Item returned to inventory and sale removed.", "success")
    except Exception as e:
        app.logger.exception("[return_sale] failed")
        flash(f"Return failed: {e}", "error")
    return redirect(url_for("index"))

# 🗑️ Delete a sale record (no restock) — PASSWORD PROTECTED
@app.post("/sales/<int:sale_id>/delete")
def delete_sale(sale_id: int):
    if not is_logged_in(): return redirect(url_for("login"))
    if not check_csrf():   return redirect(url_for("index"))
    if not require_admin_action_pw(): return redirect(url_for("index"))

    if not db.db_one("SELECT 1 FROM sales WHERE id=:id", {"id": sale_id}):
        flash("Sale record not found.", "error")
        return redirect(url_for("index"))

    db.db_exec("DELETE FROM sales WHERE id=:id", {"id": sale_id})
    _admin_log("sale_delete", f"id={sale_id}", 1)
    flash("Sale record deleted.", "info")
    return redirect(url_for("index"))

# ❌ Delete Item — PASSWORD PROTECTED
@app.post("/delete/<int:item_id>")
def delete_item(item_id):
    if not is_logged_in(): return redirect(url_for("login"))
    if not check_csrf():   return redirect(url_for("index"))
    if not require_admin_action_pw(): return redirect(url_for("index"))

    row = db.db_one("SELECT name FROM inventory WHERE id=:id", {"id": item_id})
    name = row[0] if row else "unknown"

    db.db_exec("DELETE FROM inventory WHERE id=:id", {"id": item_id})
    _admin_log("delete", name, 1)
    flash(f"Item deleted successfully — {name}!", "warning")
    return redirect(url_for("index"))


# REST alias for delete → same behavior, easier templating
@app.post("/items/<int:item_id>/delete")
def delete_item_alias(item_id: int):
    return delete_item(item_id)

# ✅ Return to stock (restock; PASSWORD-PROTECTED)
@app.post("/items/<int:item_id>/return")
def return_item(item_id: int):
    if not is_logged_in(): return redirect(url_for("login"))
    if not check_csrf():   return redirect(url_for("index"))
    if not require_admin_action_pw(): return redirect(url_for("index"))

    try:
        amt = int((request.form.get("amount") or "1").strip())
    except Exception:
        amt = 0
    if amt <= 0:
        flash("Return amount must be greater than 0.", "error")
        return redirect(url_for("index"))

    # ensure item exists
    row = db.db_one("SELECT name FROM inventory WHERE id=:id", {"id": item_id})
    if not row:
        flash("Item not found.", "error")
        return redirect(url_for("index"))
    item_name = row[0]

    db.db_exec("UPDATE inventory SET quantity = quantity + :q WHERE id=:id", {"q": amt, "id": item_id})

    # (optional but recommended) admin audit log
    try:
        _admin_log("restock_return", item_name, 1)
    except Exception:
        pass

    flash(f"Returned {amt} unit(s) to stock for “{item_name}”.", "success")

    # go back to edit if requested
    if (request.form.get("ref") or "") == "edit":
        return redirect(url_for("edit_item", item_id=item_id))
    return redirect(url_for("index"))



# ✏️ Edit Item — PASSWORD CHECK ON POST
@app.route("/edit/<int:item_id>", methods=["GET","POST"])
def edit_item(item_id: int):
    if not is_logged_in():
        flash("Please log in.", "warning")
        return redirect(url_for("login"))

    if request.method == "POST":
        if not check_csrf():
            return redirect(url_for("edit_item", item_id=item_id))
        if not require_admin_action_pw():
            return redirect(url_for("edit_item", item_id=item_id))

        # --- Read + validate fields ---
        name_raw = (request.form.get("name") or "").strip()
        if not name_raw:
            flash("Name is required.", "error")
            return redirect(url_for("edit_item", item_id=item_id))
        name = re.sub(r"\s+", " ", name_raw).strip()[:100]

        try:
            quantity = int((request.form.get("quantity") or "").strip())
        except Exception:
            flash("Quantity must be a number.", "error")
            return redirect(url_for("edit_item", item_id=item_id))
        if quantity < 0:
            flash("Quantity must be ≥ 0.", "error")
            return redirect(url_for("edit_item", item_id=item_id))

        try:
            bp_ui = Decimal(str(parse_money(request.form.get("buying_price") or "0")))
            sp_ui = Decimal(str(parse_money(request.form.get("selling_price") or "0")))
        except (InvalidOperation, Exception):
            flash("Prices must be numeric.", "error")
            return redirect(url_for("edit_item", item_id=item_id))
        if bp_ui < 0 or sp_ui < 0:
            flash("Prices must be non-negative.", "error")
            return redirect(url_for("edit_item", item_id=item_id))
        if sp_ui < bp_ui:
            flash("Warning: selling price is below buying price.", "warning")

        # Convert UI → DB (UZS)
        try:
            buying_price  = Decimal(str(convert_amount(bp_ui, from_curr=get_curr(), to_curr=BASE_CCY)))
            selling_price = Decimal(str(convert_amount(sp_ui, from_curr=get_curr(), to_curr=BASE_CCY)))
        except Exception as e:
            flash(f"Currency conversion failed: {e}", "error")
            return redirect(url_for("edit_item", item_id=item_id))

        bp_db = float(buying_price)
        sp_db = float(selling_price)

        try:
            db.db_exec(
                """
                UPDATE inventory
                   SET name=:n,
                       buying_price=:bp,
                       selling_price=:sp,
                       quantity=:q,
                       updated_at=CURRENT_TIMESTAMP
                 WHERE id=:id
                """,
                {"n": name, "bp": bp_db, "sp": sp_db, "q": quantity, "id": item_id},
            )
            _admin_log("edit", name, 1)
            flash(f"Item “{name}” updated.", "success")
            return redirect(url_for("index"))
        except Exception as e:
            app.logger.exception("[edit_item] failed")
            flash(f"Failed to update item: {e}", "error")
            return redirect(url_for("edit_item", item_id=item_id))

    # GET
    item = db.db_one("SELECT * FROM inventory WHERE id=:id", {"id": item_id})
    if not item:
        flash("Item not found.", "error")
        return redirect(url_for("index"))
    return render_template("edit.html", item=item)

# 📊 API — Stock Overview (for advanced chart)
@app.get("/api/stock/overview")
def api_stock_overview():
    if not is_logged_in():
        return jsonify({"items": [], "low_threshold": 5, "totals": {"qty": 0, "value": 0.0}}), 401

    rows = db.db_all(
        """
        SELECT id, name,
               COALESCE(quantity,0)       AS qty,
               COALESCE(buying_price,0)   AS buy,
               COALESCE(selling_price,0)  AS sell
        FROM inventory
        ORDER BY name
        """
    )

    items, totals_qty, totals_value = [], 0, 0.0
    low_threshold = 5
    ui = get_curr()

    for r in rows:
        iid, name, qty, buy, sell = int(r[0]), r[1], int(r[2]), float(r[3]), float(r[4])
        value_db = qty * sell
        value_ui = convert_amount(value_db, from_curr=BASE_CCY, to_curr=ui)
        items.append({
            "id": iid, "name": name, "qty": qty,
            "buy": convert_amount(buy,  from_curr=BASE_CCY, to_curr=ui),
            "sell": convert_amount(sell, from_curr=BASE_CCY, to_curr=ui),
            "value": value_ui,
            "profit_per": max(convert_amount(sell - buy, from_curr=BASE_CCY, to_curr=ui), 0),
            "is_low": qty <= low_threshold,
        })
        totals_qty   += qty
        totals_value += value_ui

    return jsonify({
        "items": items,
        "low_threshold": low_threshold,
        "totals": {"qty": totals_qty, "value": totals_value},
    })

# 📊 Export to Excel (dual currency)
@app.get("/export")
def export_excel():
    if not is_logged_in():
        return redirect(url_for("login"))

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule

    ui_curr = get_curr()
    base_ccy = BASE_CCY
    inventory = get_inventory()

    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = "Inventory"

    sh["A1"] = "Vanta Inventory Export"
    sh["A2"] = f"Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    sh["A3"] = f"Stored currency (DB): {base_ccy}"
    sh["A4"] = f"UI currency: {ui_curr}"

    start_row = 6
    headers = [
        "ID", "Name",
        f"Buying ({base_ccy})", f"Selling ({base_ccy})", "Quantity", f"Profit ({base_ccy})",
        f"Buying ({ui_curr})",  f"Selling ({ui_curr})",  f"Profit ({ui_curr})",
    ]

    for c, h in enumerate(headers, start=1):
        cell = sh.cell(row=start_row, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sh.freeze_panes = f"A{start_row+1}"
    sh.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row}"

    fmt_int      = '#,##0'
    fmt_ccy_base = f'#,##0 "{base_ccy}"'
    fmt_ccy_ui   = f'#,##0 "{ui_curr}"'

    rowi = start_row + 1
    first_row = rowi
    for it in inventory:
        _id, name, buy_uzs, sell_uzs, qty, profit_uzs, _curr = it
        buy_ui  = convert_amount(buy_uzs,  from_curr=base_ccy, to_curr=ui_curr)
        sell_ui = convert_amount(sell_uzs, from_curr=base_ccy, to_curr=ui_curr)
        prof_ui = convert_amount(profit_uzs, from_curr=base_ccy, to_curr=ui_curr)

        sh.cell(row=rowi, column=1, value=int(_id))
        sh.cell(row=rowi, column=2, value=name)

        c3 = sh.cell(row=rowi, column=3, value=float(buy_uzs or 0));    c3.number_format = fmt_ccy_base
        c4 = sh.cell(row=rowi, column=4, value=float(sell_uzs or 0));   c4.number_format = fmt_ccy_base
        c5 = sh.cell(row=rowi, column=5, value=int(qty or 0));          c5.number_format = fmt_int
        c6 = sh.cell(row=rowi, column=6, value=float(profit_uzs or 0)); c6.number_format = fmt_ccy_base

        c7 = sh.cell(row=rowi, column=7, value=float(buy_ui or 0));     c7.number_format = fmt_ccy_ui
        c8 = sh.cell(row=rowi, column=8, value=float(sell_ui or 0));    c8.number_format = fmt_ccy_ui
        c9 = sh.cell(row=rowi, column=9, value=float(prof_ui or 0));    c9.number_format = fmt_ccy_ui

        rowi += 1

    last_row = rowi - 1

    sh.cell(row=rowi, column=2, value="TOTALS").font = Font(bold=True)
    t_qty   = sh.cell(row=rowi, column=5, value=f"=SUM(E{first_row}:E{last_row})");  t_qty.number_format   = fmt_int
    t_pB    = sh.cell(row=rowi, column=6, value=f"=SUM(F{first_row}:F{last_row})");  t_pB.number_format    = fmt_ccy_base
    t_pU    = sh.cell(row=rowi, column=9, value=f"=SUM(I{first_row}:I{last_row})");  t_pU.number_format    = fmt_ccy_ui

    thin = Side(style="thin", color="DDDDDD")
    rng = sh[f"A{start_row}:I{rowi}"]
    for rw in rng:
        for c in rw:
            c.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    sh.conditional_formatting.add(
        f"E{first_row}:E{last_row}",
        CellIsRule(operator="lessThanOrEqual", formula=["5"],
                   fill=PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"))
    )

    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        maxlen = 0
        for r in range(1, rowi + 1):
            v = sh.cell(row=r, column=col).value
            vlen = len(str(v)) if v is not None else 0
            maxlen = max(maxlen, vlen)
        sh.column_dimensions[letter].width = min(max(10, maxlen + 2), 42)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"inventory_export_{ui_curr}_and_{base_ccy}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# 📦 Data backup (ZIP of CSVs, Postgres)
@app.get("/admin/backup")
def admin_backup():
    if not (is_logged_in() and is_admin_user()):
        return redirect(url_for("login"))

    db_url = os.getenv("DATABASE_URL")
    if not db_url:
        return "DATABASE_URL not set", 500

    try:
        import psycopg2
    except Exception as e:
        return f"psycopg2 not available on this environment: {e}", 500

    if db_url.startswith("postgres://"):
        db_url = db_url.replace("postgres://", "postgresql://", 1)
    parsed = urlparse(db_url)
    user = parsed.username
    pwd = parsed.password
    host = parsed.hostname
    port = parsed.port or 5432
    dbname = parsed.path.strip("/")

    conn = psycopg2.connect(dbname=dbname, user=user, password=pwd, host=host, port=port)
    conn.autocommit = True

    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT table_name
            FROM information_schema.tables
            WHERE table_schema='public' AND table_type='BASE TABLE'
            ORDER BY table_name;
            """
        )
        tables = [r[0] for r in cur.fetchall()]

    ts = datetime.utcnow().strftime("%Y-%m-%dT%H-%M-%SZ")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        meta = {
            "generated_at_utc": ts,
            "database": dbname,
            "host": host,
            "tables": tables,
            "note": "Data-only backup (CSV per table). Schema ensured at startup.",
        }
        zf.writestr("metadata.json", json.dumps(meta, indent=2))
        for tname in tables:
            with conn.cursor() as c2:
                s = io.StringIO()
                c2.copy_expert(f'COPY (SELECT * FROM "{tname}") TO STDOUT WITH CSV HEADER', s)
                zf.writestr(f"{tname}.csv", s.getvalue())

    conn.close()
    buf.seek(0)
    fname = f"vanta_inventory_backup_{ts}.zip"
    return send_file(buf, mimetype="application/zip", as_attachment=True, download_name=fname)

# =========================
# Prefs & small APIs
# =========================
@app.get("/api/rates")
def api_rates():
    try:
        base = (request.args.get("base") or get_curr()).upper()
        base, rates = _derive_rates_from_usd(base)
        return jsonify({"base": base, "rates": rates})
    except Exception as e:
        return jsonify({
            "base": "USD",
            "rates": {"USD": 1.0, "AED": 3.6725, "UZS": 12600.0},
            "_note": "fallback",
            "_error": str(e)
        })

@app.get("/api/geo")
def api_geo():
    if OFFLINE:
        return jsonify({"country": "US", "languages": "en", "_note": "offline default"})

    try:
        r = requests.get("https://ipapi.co/json", timeout=6)
        r.raise_for_status()
        j = r.json()
        if not isinstance(j, dict):
            raise ValueError("bad geo json")
        return jsonify(j)
    except Exception as e:
        return jsonify({"country": "US", "languages": "en", "_note": "fallback", "_error": str(e)})

@app.post("/prefs")
def set_prefs():
    if not check_csrf():
        return redirect(url_for("login"))
    # Only set what was actually submitted
    lang = (request.form.get("lang") or "").strip()
    if lang:
        session["LANG"] = "uz" if lang == "uz" else "en"
    curr = (request.form.get("curr") or "").strip().upper()
    if curr and curr in _SUPPORTED:
        session["CURR"] = curr
    flash(t("preferences_saved", session.get("LANG", DEFAULT_LANG)), "success")
    return redirect(url_for("login"))

@app.post("/settings/currency")
def set_currency():
    if not is_logged_in():
        return redirect(url_for("login"))
    if not check_csrf():
        return redirect(url_for("index"))
    cur = (request.form.get("currency") or DEFAULT_CURR).upper()
    if cur not in _SUPPORTED:
        flash("Unsupported currency.", "error")
        return redirect(url_for("index"))
    session["CURR"] = cur
    flash(f"Currency set to {cur}.", "success")
    return redirect(url_for("index"))

# =========================
# Auth pages (hardened)
# =========================
@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("logged_in"):
        flash("You are already logged in.", "info")
        return redirect(url_for("index"))

    if request.method == "POST":
        if not check_csrf():
            return redirect(url_for("login"))

        username = (request.form.get("username") or "").strip().casefold()
        password = (request.form.get("password") or "").strip()

        stored = (ADMIN_USERS.get(username) or "")
        ok = hmac.compare_digest(stored, password)

        if ok:
            session["logged_in"] = True
            session["user"] = username
            flash("Welcome back.", "success")
            return redirect(url_for("index"))

        if app.debug:
            app.logger.info(f"[login] failed for user='{username}'  known={list(ADMIN_USERS.keys())}")
        flash("Invalid username or password.", "error")
        return redirect(url_for("login"))

    return render_template("login.html", APP_VERSION=get_version())

@app.get("/logout")
def logout():
    user = session.get("user")
    session.clear()
    if app.debug and user:
        app.logger.info(f"[logout] user={user} logged out")
    flash("You’ve been logged out.", "info")
    return redirect(url_for("login"))

@app.get("/forgot")
def forgot_password():
    return render_template("forgot.html", APP_VERSION=get_version())

@app.post("/forgot")
def forgot_password_post():
    if not check_csrf():
        return redirect(url_for("forgot_password"))
    email = (request.form.get("email") or "").strip()
    if not email:
        flash("Please enter your email.", "error")
        return redirect(url_for("forgot_password"))
    # TODO: Implement email reset flow
    flash("If that email exists, we’ve sent reset instructions.", "success")
    return redirect(url_for("login"))

# =========================
# Debug helpers
# =========================
@app.get("/__debug/now")
def __debug_now():
    using_sqlite = db.DATABASE_URL.startswith("sqlite")
    server_now = datetime.utcnow().isoformat() + "Z"
    if using_sqlite:
        r = db.db_one("SELECT datetime('now'), date('now')")
        return jsonify({"server_utc": server_now, "sqlite_datetime_now": r[0], "sqlite_date_now": r[1]})
    else:
        r = db.db_one("SELECT NOW() AT TIME ZONE 'UTC', CURRENT_DATE")
        return jsonify({"server_utc": server_now, "pg_now_utc": str(r[0]), "pg_current_date": str(r[1])})

@app.get("/__debug/sales")
def __debug_sales():
    rows = db.db_all("SELECT id, item_id, qty, sell_price, profit, sold_at FROM sales ORDER BY id DESC LIMIT 10")
    return jsonify([{ 
        "id": r[0], "item_id": r[1], "qty": r[2], "sell_price": float(r[3] or 0),
        "profit": float(r[4] or 0), "sold_at": str(r[5])
    } for r in rows])

@app.post("/__admin/wipe")
def __admin_wipe():
    if not (app.debug and is_logged_in()):
        return "Forbidden", 403
    if not check_csrf():
        return redirect(url_for("index"))
    try:
        db.db_exec("DELETE FROM sales")
        db.db_exec("DELETE FROM inventory")
        if db.DATABASE_URL.startswith("sqlite"):
            db.db_exec("VACUUM")
        flash("Local DB wiped.", "success")
    except Exception as e:
        flash(f"Failed to wipe DB: {e}", "error")
    return redirect(url_for("index"))

@app.get("/__debug_sales_today")
def __debug_sales_today():
    using_sqlite = db.DATABASE_URL.startswith("sqlite")
    if using_sqlite:
        q = """
        SELECT COUNT(*), COALESCE(SUM(qty*sell_price),0), COALESCE(SUM(profit),0)
        FROM sales
        WHERE sold_at >= DATETIME('now','start of day')
          AND sold_at <  DATETIME('now','start of day','+1 day')
        """
    else:
        q = """
        SELECT COUNT(*), COALESCE(SUM(qty*sell_price),0), COALESCE(SUM(profit),0)
        FROM sales
        WHERE sold_at >= CURRENT_DATE
          AND sold_at <  CURRENT_DATE + INTERVAL '1 day'
        """
    row = db.db_one(q)
    return {"rows": row[0], "revenue_base": float(row[1] or 0), "profit_base": float(row[2] or 0)}

@app.get("/test-flash")
def test_flash():
    flash_success("Success message")
    flash_info("Info message")
    flash_warn("Warning message")
    flash_error("Error message")
    return redirect(url_for("index"))

# =========================
# Admin Logs page
# =========================
@app.get("/admin/logs")
def admin_logs():
    if not is_logged_in():
        return redirect(url_for("login"))
    rows = db.db_all("SELECT id, action, item, success, ts FROM admin_logs ORDER BY ts DESC LIMIT 100")
    logs = [{
        "id": r[0],
        "action": r[1],
        "item": r[2],
        "success": bool(r[3]),
        "ts": str(r[4])
    } for r in rows]
    return render_template("admin_logs.html", logs=logs)

# =========================
# Dev entry
# =========================
if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=(ENV=="dev"))
